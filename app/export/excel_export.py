"""Geração da planilha Excel de importação para o Domínio Sistemas.

O layout inicial segue o especificado (Data, Conta Débito, Conta Crédito,
Histórico, Documento, Valor, Centro de Custo, Filial). Caso o Domínio
Sistemas exija um modelo Excel/TXT específico do cliente, ajuste
`COLUNAS_EXPORTACAO` e `_montar_linha` — o restante do pipeline não muda.
"""
from __future__ import annotations

import io
import logging

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import ClassificacaoResultado, TipoLancamento

logger = logging.getLogger(__name__)

COLUNAS_EXPORTACAO = [
    "Data", "Conta Débito", "Conta Crédito", "Histórico",
    "Documento", "Valor", "Centro de Custo", "Filial",
]


def build_export_dataframe(
    resultados: list[ClassificacaoResultado],
    conta_banco_codigo: str,
    centro_custo: str = "",
    filial: str = "",
) -> pd.DataFrame:
    linhas = [
        _montar_linha(r, conta_banco_codigo, centro_custo, filial) for r in resultados
    ]
    return pd.DataFrame(linhas, columns=COLUNAS_EXPORTACAO)


def _montar_linha(
    resultado: ClassificacaoResultado, conta_banco_codigo: str, centro_custo: str, filial: str
) -> dict:
    lancamento = resultado.lancamento
    if lancamento.tipo == TipoLancamento.CREDITO:
        # Dinheiro entrou no banco: debita o banco, credita a conta classificada.
        conta_debito, conta_credito = conta_banco_codigo, resultado.conta_codigo
    else:
        # Dinheiro saiu do banco: debita a conta classificada, credita o banco.
        conta_debito, conta_credito = resultado.conta_codigo, conta_banco_codigo

    return {
        "Data": lancamento.data.strftime("%d/%m/%Y"),
        "Conta Débito": conta_debito,
        "Conta Crédito": conta_credito,
        "Histórico": lancamento.historico,
        "Documento": lancamento.documento,
        "Valor": lancamento.valor,
        "Centro de Custo": centro_custo,
        "Filial": filial,
    }


def export_to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Importacao Dominio")
        _estilizar_planilha(writer.sheets["Importacao Dominio"], df)
    logger.info("Planilha de exportação gerada com %d linhas", len(df))
    return buffer.getvalue()


def _estilizar_planilha(worksheet, df: pd.DataFrame) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        largura = max(12, min(40, int(df[column].astype(str).str.len().max() or 0) + 4))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = largura
